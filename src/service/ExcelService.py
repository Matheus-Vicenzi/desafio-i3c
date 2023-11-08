import openpyxl
import traceback
from infra.Enviroment import EXCEL_FILE_PATH


def change_item(item_id, new_quantity_value):

    try:
        print("change item excel inicio")
        cell_id_bom = get_item_cell_id(item_id)
        set_item_quantity(cell_id_bom, new_quantity_value)
        proposta_row = get_item_proposta_row(cell_id_bom)
        total_value = get_item_total_value(proposta_row)
        print("change item excel fim. total value: " + str(total_value))
        return total_value
    
    except Exception as e:
        print(e)
        traceback.print_exc()
        raise e


def set_item_quantity(cell_id_bom, new_quantity_value):
    try:
        wb = openpyxl.load_workbook(
            EXCEL_FILE_PATH, data_only=False)
        sheet = wb['BOM']

        quantity_cell = sheet.cell(row=cell_id_bom.row, column=9)
        quantity_cell.value = new_quantity_value
        

        wb.save(EXCEL_FILE_PATH)

    except Exception as e:
        print(e)
        traceback.print_exc()

    finally:
        wb.close()


def get_item_proposta_row(cell_id_bom):
    try:
        wb = openpyxl.load_workbook(
            EXCEL_FILE_PATH, data_only=False)
        sheet = wb['PROPOSTA']

        id_column = sheet['B']
        for cell in id_column:
            print(cell.value)
            if cell_id_bom.coordinate in str(cell.value):
                print(cell.row)
                return cell.row

    except Exception as e:
        print(e)
        traceback.print_exc()

    finally:
        wb.close()


def get_item_cell_id(item_id):
    try:
        wb = openpyxl.load_workbook(
            EXCEL_FILE_PATH, data_only=False)
        sheet = wb['BOM']

        id_column = sheet['A']
        for cell in id_column:
            if cell.value == None:
                continue
            print(cell.value)
            # print("=A" in str(cell.value))
            # if "=A" in str(cell.value):
            #     print("entrou")
            #     cell_value = str(cell.value)
            #     cell_value = cell_value.replace("=A", "")
            #     cell_value = cell_value.replace("+1", "")
            #     print(cell_value)
            #     cell_value = int(cell_value) + 1 - 4
            #     if cell_value > 3:
            #         cell_value = cell_value + 1
            # if cell_value == item_id:
            if cell.value == item_id:
                print(cell.coordinate)
                return cell
            
        raise Exception("Item não encontrado")

    except Exception as e:
        print(e)
        traceback.print_exc()
        raise e

    finally:
        wb.close()


def change_item_price(sheet, id_cell, new_price_value):
    try:
        value_cell = sheet.cell(row=id_cell.row, column=10)
        value_cell.value = new_price_value
    except Exception as e:
        print(e)
        traceback.print_exc()


def change_item_quantity(sheet, id_cell, new_quantity_value):
    try:
        quantity_cell = sheet.cell(row=id_cell.row, column=9)
        quantity_cell.value = new_quantity_value
    except Exception as e:
        print(e)
        traceback.print_exc()


def get_item_total_value(proposta_row):
    try:
        wb = openpyxl.load_workbook(
            EXCEL_FILE_PATH, data_only=True)
        
        sheet = wb['PROPOSTA']

        cell = sheet.cell(row=proposta_row, column=9)
        print(cell.value)

        return cell.value

    except Exception as e:
        print(e)
        traceback.print_exc()

    finally:
        wb.close()
